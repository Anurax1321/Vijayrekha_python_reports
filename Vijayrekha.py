## FILE HEADER
## Program Desciption

## Imports
import os
import sys
import pandas as pd
import numpy as np
import re
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill

# Declaration

actual_no_of_patients = 0
name = []
compound = []
response = []
final_result = []


# Function to highlight the cells based on thier values
def cell_highlight(min, max, cell, b, y, g):
    ## TODO: Check for the correct inputs if needed
    try:
        if not min <= cell.value <= max:
            if not cell.value <= max:
                cell.font = b
            cell.fill = y
        else:
            cell.fill = g
    except Exception as e:
        print(f"Problem in Cell_Highlight method; Check it out!!! {e}")
        sys.exit(1)


# Function to Highlight the cells of the excel based on the range
def data_outliners(final_data_frame, excel_path):
    print("Checking for outliners; ", end="")
    try:
        wb = load_workbook(excel_path)
        ws = wb.active

        range_dict = {'Ala':(103,742), 'Arg':(1,41), 'Asp':(10,345), 'Cit':(5,43), 'Glu':(152,708), 'Gly':(0,1142),
                      'Leu':(27,324), 'Met':(5,41), 'Orn':(10,263), 'Phe':(10,102), 'Pro':(87,441), 'Tyr':(15,259),
                      'Val':(52,322), 'C0':(5,125), 'C2':(1.4,80), 'C3':(0.18,0.63), 'C4':(0.08,1.7), 'C5':(0.01,1),
                      'C5DC':(0.01,2.99), 'C6':(0.01,0.95), 'C8':(0.01,0.6), 'C10':(0.02,0.65), 'C12':(0.02,0.6),
                      'C14':(0.01,1.22), 'C16':(0.34,10.35), 'C18':(0.21,2.03), 'C5:1':(0.01,0.9), 'C4OH':(0.01,1.29),
                      'C5OH':(0.01,0.9), 'C8:1':(0.01,0.7), 'C3DC':(0.1,0.45), 'C10:2':(0.01,0.22), 'C10:1':(0.01,0.45),
                      'C4DC':(0.1,1.25), 'C12:1':(0.01,0.5), 'C6DC':(0.01,0.23), 'C14:2':(0,0.2), 'C14:1':(0.01,0.8),
                      'C14OH':(0,0.2), 'C16:1':(0.01,1.4), 'C16:1OH':(0.01,0.1), 'C16OH':(0.01,0.1), 'C18:2':(0.1,0.73),
                      'C18:1':(0.5,7), 'C18:2OH':(0.01,0.03), 'C18:1OH':(0.01,0.1), 'C18OH':(0.01,0.1)}

        col_1_range_dict = {'Ala':(338,702), 'Arg':(6.07,12.6), 'Asp':(17.2,45.5), 'Cit':(15.8,41.6), 'Glu':(392,813),
                            'Gly':(301,701), 'Leu':(125,259), 'Met':(14.7,34.2), 'Orn':(101,237), 'Phe':(67.7,126),
                            'Pro':(174,323), 'Tyr':(47.1,110), 'Val':(108,284), 'C0':(14.3,57.1), 'C2':(3.56,15.8),
                            'C3':(0.326,0.861), 'C4':(0.13,0.27), 'C5':(0.262,0.611), 'C5DC':(0.487,1.95),
                            'C6':(0.072,0.189), 'C8':(0.237,0.624), 'C10':(0.12,0.318), 'C12':(0.27,0.561),
                            'C14':(0.279,0.652), 'C16':(0.869,2.03), 'C18':(0.425,0.991)}

        col_2_range_dict = {'Ala':(731,1519), 'Arg':(12.4,28.9), 'Asp':(157,413), 'Cit':(135,404), 'Glu':(536,995),
                            'Gly':(755,1568), 'Leu':(375,876), 'Met':(231,479), 'Orn':(361,843), 'Phe':(471,979),
                            'Pro':(398,828), 'Tyr':(359,745), 'Val':(325,757), 'C0':(68.1,321), 'C2':(5.2,20.8),
                            'C3':(5.51,11.4), 'C4':(0.436,0.809), 'C5':(1.26,2.62), 'C5DC':(1.78,7.1),
                            'C6':(0.506,1.05), 'C8':(1.56,3.24), 'C10':(0.663,1.38), 'C12':(3.81,7.07),
                            'C14':(2.0,4.16), 'C16':(6.98,14.5), 'C18':(2.79,5.8)}

        bold = Font(bold=True)

        green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        for col in final_data_frame.columns:
            if not col == "Name":
                i = 0
                assert col in range_dict, "Unkown column in the excel; Check the range for errors"
                col_idx = final_data_frame.columns.get_loc(col)+1
                col_letter = ws.cell(row = 1, column = col_idx).column_letter

                for cell in ws[col_letter][1:]:
                    if isinstance(cell.value, (int,float)):
                        if i < 2:
                            if col in col_1_range_dict:
                                min_value, max_value = col_1_range_dict[col]
                                cell_highlight(min_value, max_value, cell, bold, yellow_fill, green_fill)
                                i += 1
                            else:
                                i += 1
                                continue
                        elif i < 4:
                            if col in col_2_range_dict:
                                min_value, max_value = col_2_range_dict[col]
                                cell_highlight(min_value, max_value, cell, bold, yellow_fill, green_fill)
                                i += 1
                            else:
                                i += 1
                                continue
                        else:
                            assert col in range_dict, "Column not in the range Dictionary!!; Problem"
                            min_value, max_value = range_dict[col]
                            cell_highlight(min_value, max_value, cell, bold, yellow_fill, green_fill)
    except Exception as e:
        print(f"Probelm with the highlighting. \n{e}")
        sys.exit(1)
    print("Highlighting Completed.")
    try:
        wb.save(excel_path)
        print("Successfully saves to the excel")
    except Exception as e:
        print(f"Problem with the saving. \n{e}")
        sys.exit(1)


## Function to find the increment number in the file path
def increment_match(match):
    return f"({int(match.group(1)) + 1})"


## Fucntion to write to a file
def write_to_excel(final_df, excel_path):
    try:
        i = 1
        while True:
            if not os.path.isfile(excel_path):
                final_df.to_excel(excel_path, index = False)
                break

            print("File Already exist with the same name; So fetching a different file path with the same date")
            print("\nTry not doing this again!!\n")
            if i == 1:
                date = excel_path.split(".x")
                excel_path = date[0] + "(1).x" + date[1]
                i += 1
                continue
            else:
                pattern = r"\((\d+)\)"
                assert re.match(pattern,excel_path), "Pattern not found in the file path!!!"
                excel_path = re.sub(pattern, increment_match, excel_path)
                continue
        print(f"Excel File Path is: {excel_path}")
        print("Writing to the Excel File; ", end="",flush=True)
        return excel_path
    except AssertionError as e:
        print(f"Problem in the Excel File Path\n{e}")
        sys.exit(1)
    except Exception as e:
        print(f"Writing to the Excel File Unsuccessfull\n{e}")
        sys.exit(1)

## Function to extract the data and create the data-frame into which we can perform some manupilations
## as we require and also reshape the list to a 2D Array for the dataframe
## Returns the final Data Frame for each file type; There are three File types- AA;AC;AC_EXT
def data_extraction(filePath):
    no_of_patients = 0
    if os.path.exists(filePath):
        if os.path.isfile(filePath) and filePath.endswith(".txt"):
            try:
                with open(filePath, 'r') as file:
                    for line in file:
                        if line == '\n':
                            continue
                        if line.strip().startswith("Compound"):

                            ## Breaking Section; Stopping mechanism
                            if filePath.endswith("_AA.txt" ): ## AA Data
                                if "Suac" in line:
                                    break
                            elif filePath.endswith("_AC.txt"): ## AC Data
                                if line.strip().endswith("IS"):
                                    break
                            elif filePath.endswith( "_AC_EXT.txt"): ## AC_EXT Data
                                if line.strip().endswith("IS"):
                                    break
                            else:
                                print("You should not see this message.\nSomething Wrong with file type and Breaking\nRefer break section in Data Extraction.")

                            ## Splitting and Stripping the coumpound name and then adding it to the list
                            line = line.split('  ')
                            compound.append(line[1].strip())
                            continue
                        elif line.strip()[0].isdigit(): ## Lines starting with a number
                            line = line.split('\t')
                            if no_of_patients < int(line[0]):
                                no_of_patients = int(line[0]) # Assigning the maximum number which is the no. of patients
                            # Extracting the name
                            if line[1] not in name:
                                name.append(line[1])
                            # the last element is the parameter of interest
                            response.append(line[-1].strip())
                        else:
                            continue # any other unwanted lines just pass through

                    # Checking the extraction
                    if actual_no_of_patients != no_of_patients:
                        print("Issue with the Number of patients;\nCheck the value entered and the file!! Try Again")
                        sys.exit(1)  ## Exciting the program with error code '1'

                    # ## Printing for checking the extraction
                    print(compound)
                    # print(name)
                    # print(response)
                    # print(no_of_patients)

                    # Data Extraction Successful
                    print("Data Extraction Complete !!")
                    print("Creating the Final Results")

            except FileNotFoundError as e:
                print("File is not found in the right place\nPut the file in the right place or give the right file path Please!!!\n--Anurag:))\n{e}")
                sys.exit(1)
            except UnboundLocalError as e:
                print(f"Local Variable being used before assignment!!\n{e}\nContact Anurag ")
                sys.exit(1)
            except Exception as e:
                print(f"Problem with Something; Check it out!!! {e}")
                sys.exit(1)

        else:
            if filePath.endswith("_AA.txt"):  ## AA Data
                print("\nAA File is invalid; Try Again!!")
                sys.exit(1)
            elif filePath.endswith("_AC.txt"):  ## AC Data
                print("\nAC File is invalid; Try Again!!")
                sys.exit(1)
            elif filePath.endswith("_AC_EXT.txt"):  ## AC_EXT Data
                print("\nAC_EXT File is invalid; Try Again!!")
                sys.exit(1)
            else:
                print("\nFile is invalid and not detected")
                sys.exit(1)
    else:
        if filePath.endswith("_AA.txt"):  ## AA Data
            print("\nAA File Path is invalid; Try Again!!")
            sys.exit(1)
        elif filePath.endswith("_AC.txt"):  ## AC Data
            print("\nAC File Path is invalid; Try Again!!")
            sys.exit(1)
        elif filePath.endswith("_AC_EXT.txt"):  ## AC_EXT Data
            print("\nAC_EXT File Path is invalid; Try Again!!")
            sys.exit(1)
        else:
            print("\nFile Path is invalid and not detected")
            sys.exit(1)

    ## Manupilate the data

    # Counter Variables
    k = 0  # Counts the total number of values
    count = 0  # Counts the total number of compounds

    ## Multiply with the right number based on the file path
    for i in compound:
        count += 1  # Number of Compounds
        j = 0
        while actual_no_of_patients > j:

            if k >= len(response):
                print("Not enough responses for the given number of patients and compounds.")
                sys.exit(1)

            if response[k] == '':
                final_result.append('0.0')
                k += 1
                j += 1
                continue

            if filePath.endswith("_AA.txt" ): ## AA Data
                if i == "Gly":
                    result = float(response[k]) * 403
                    final_result.append(result)
                elif i != "Gly":
                    result = float(response[k]) * 80.6
                    final_result.append(result)
                else:
                    print("AA file; Some issue with compound checking and Data manipulation")
                    sys.exit(1)
            elif filePath.endswith("_AC.txt" ): ## AC Data
                if i == "C0":
                    result = float(response[k]) * 20.8
                    final_result.append(result)
                elif i == "C2":
                    result = float(response[k]) * 6.13
                    final_result.append(result)
                elif i in ["C3", "C4", "C8"]:
                    result = float(response[k]) * 1.29
                    final_result.append(result)
                elif i == "C5DC":
                    result = float(response[k]) * 1.48
                    final_result.append(result)
                elif i == "C14":
                    result = float(response[k]) * 0.903
                    final_result.append(result)
                elif i in ["C16", "C18"]:
                    result = float(response[k]) * 3.35
                    final_result.append(result)
                elif i in ["C5","C6","C10","C12"]:
                    result = float(response[k]) * 1.1
                    final_result.append(result)
                else:
                    print("AC file; Some issue with compound checking and Data manipulation")
                    sys.exit(1)
            elif filePath.endswith("_AC_EXT.txt" ): ## AC_EXT Data
                if i in ["C5:1", "C5OH", "C10:2", "C10:1","C12:1", "C6DC"]:
                    result = float(response[k]) * 1.1
                    final_result.append(result)
                elif i in ["C4OH", "C8:1", "C3DC", "C4DC"]:
                    result = float(response[k]) * 1.29
                    final_result.append(result)
                elif i in ["C14:2", "C14:1", "C14OH"]:
                    result = float(response[k]) * 0.903
                    final_result.append(result)
                elif i in ["C16:1","C16:1OH","C16OH","C18:2","C18:1","C18:2OH","C18:1OH","C18OH"]:
                    result = float(response[k]) * 3.35
                    final_result.append(result)
                else:
                    print("AC_EXT file; Some issue with compound checking and Data manipulation")
                    sys.exit(1)
            else:
                print("Wrong file type; Data Manipulation Section")
                sys.exit(1)

            # Updating the indices
            k+=1
            j+=1
    if not len(response) == len(final_result):
        print("Wrong result; Change the logic for Manupilation")
        sys.exit(1)

    ## Reshape and storing returning the individual data file

    print("Reshaping the final results")
    if filePath.endswith("_AA.txt"):  ## AA Data
        AA_2d_array = np.array(final_result).reshape(len(name),len(compound),order='F')
        print("AA Data Extraction and manipulation complete\nStoring the Data Frame")
        return pd.DataFrame(AA_2d_array, columns=compound)
    elif filePath.endswith("_AC.txt"):  ## AC Data
        AC_2d_array = np.array(final_result).reshape(len(name),len(compound),order='F')
        print("AC Data Extraction and manipulation complete\nStoring the Data Frame")
        return pd.DataFrame(AC_2d_array, columns=compound)
    elif filePath.endswith("_AC_EXT.txt"):  ## AC_EXT Data
        AC_EXT_2d_array = np.array(final_result).reshape(len(name), len(compound), order='F')
        print("AC_EXT Data Extraction and manipulation complete\nStoring the Data Frame")
        return pd.DataFrame(AC_EXT_2d_array, columns=compound)
    else:
        print("Wrong file path; Data Manipulation")
        sys.exit(1)


## Function to check the inputs for their dates and formates
def check_input(filePath, date):
    filePath = filePath.split("\\")[-1]
    if filePath.endswith("_AA.txt"):
        assert filePath.startswith(date), "Wrong AA DATE!!; Check the File Paths for Data Sets and Try again!!"
    elif filePath.endswith("_AC.txt"):
        assert filePath.startswith(date), "Wrong AC DATE!!; Check the File Paths for Data Sets and Try again!!"
    elif filePath.endswith("_AC_EXT.txt"):
        assert filePath.startswith(date), "Wrong AC_EXT DATE!!; Check the File Paths for Data Sets and Try again!!"
    else:
        # Wrong Data Set Type or Formate
        return False
    return True


# TODO: Add verification using the dates form the file path in that formate before releasing it
# Function to take the file path from the user
def get_path():
    print("Requesting the File Paths for the Data Sets...")
    try:
        AA_data = input("Enter the file path for AA Data\n")
        date = get_date(AA_data)
        assert check_input(AA_data, date)
        AC_data = input("Enter the file path for AC Data\n")
        assert check_input(AC_data, date)
        ACEXT_data = input("Enter the file path for ACEXT Data\n")
        assert check_input(ACEXT_data, date)
    except AssertionError as e:
        print(f"Wrong Data File Path Entered; Try again with the correct one!!!\n{e}")
        sys.exit(1)
    except Exception as e:
        print(f"Problem in file path input; Check it out!!!\n{e}")
        sys.exit(1)
    print("Thank you for your inputs; File Paths Loaded Successfully")
    return [AA_data,AC_data,ACEXT_data]


# Function to get the date of the report
def get_date(filePath):
    date = os.path.basename(filePath).split("_")[0].strip()
    return date


# Function to concatenate all the data frames
def get_final_data(AA, AC, AC_EXT, filePath):
    final = pd.concat( [AA,AC,AC_EXT] , axis = 1)
    final.insert(0, 'Name', name)

    for idx,row in final.iterrows():
        if idx < 4:
            if row["Name"].startswith(get_date(filePath) + "_Recipe_0"):
                row['Name'] = f"CONTROL {row.name +1}"
                final.loc[idx] = row
            else:
                print("Something wrong with the alignment of the first four controls.\nProblem when renaming the controls.\nCheck get_final_date().")
                sys.exit(1)
        else:
            break

    for col in final.columns:
        if col != 'Name':
            final[col] = pd.to_numeric(final[col])
    print(final)
    return final


## Main method to exceute the program
if __name__ == '__main__':

    # default initialization
    final_data_frame = None
    AA_data_frame = None
    AC_data_frame = None
    AC_EXT_data_frame = None
    actual_no_of_patients = int(input("Enter the Number of Patients (Excluding the controls): ")) + 4
    paths = get_path()
    print("Data Extraction...")

    # TODO: CHeck for dates in the files; from the path.

    for filePath in paths:

        ### Default Initializations ###
        name = []
        compound = []
        response = []
        final_result = []

        if filePath.endswith("_AA.txt"):  ## AA Data
            print("Loading AA File")
            AA_data_frame = data_extraction(filePath)
            #print(AA_data_frame)
        elif filePath.endswith("_AC.txt"):  ## AC Data
            print("Loading AC File")
            AC_data_frame = data_extraction(filePath)
            #print(AC_data_frame)
        elif filePath.endswith("_AC_EXT.txt"):  ## AC_EXT Data
            print("Loading AC_EXT File")
            AC_EXT_data_frame = data_extraction(filePath)
            #print(AC_EXT_data_frame)
        else:
            print("Invalid Path\nPlease Try Again!")
            sys.exit(1)

    print("Data Extraction Complete for all files\nCreating a Final Data Frame for Concatenation of all the data.")
    final_data_frame = get_final_data(AA_data_frame, AC_data_frame, AC_EXT_data_frame, paths[0])
    data_outliners(final_data_frame, write_to_excel(final_data_frame,'VASU\Final Result'+ '\\' + get_date(paths[0]) + "_finalReport.xlsx"))
    print("CONGRATULATIONS!!!\nReport ready to be viewed in Excel Formate\nThank you for using the services.")


